"""Extractor de Excel (.xlsx)."""
from __future__ import annotations

from pathlib import Path

from ..models import ExtractedDoc
from .base import ExtractionError, Extractor


class XlsxExtractor(Extractor):
    extensions = (".xlsx",)

    def extract(self, path: Path) -> ExtractedDoc:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise ExtractionError("Falta dependencia: openpyxl") from exc

        wb = load_workbook(str(path), read_only=True, data_only=True)
        blocks: list[str] = []
        for ws in wb.worksheets:
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if cells:
                    rows.append(" | ".join(cells))
            if rows:
                blocks.append(f"# Hoja: {ws.title}\n" + "\n".join(rows))
        wb.close()

        meta = {"format": "xlsx", "sheets": wb.sheetnames}
        return ExtractedDoc(text="\n\n".join(blocks), meta=meta)
